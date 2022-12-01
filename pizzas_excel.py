import pandas as pd
import datetime
import re
import warnings
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference


warnings.filterwarnings("ignore")


def compilar_patrones():
    espacio = re.compile(r'\s')
    guion = re.compile(r'-')
    arroba = re.compile(r'@')
    d_0 = re.compile(r'0')
    d_3 = re.compile(r'3')
    uno = re.compile(r'one', re.I)
    dos = re.compile(r'two', re.I)
    comma = re.compile(r',')
    espacio = re.compile(r'\s')

    quitar = [espacio, guion, arroba, d_0, d_3, uno, dos]
    poner = ['_', '_', 'a', 'o', 'e', '1', '2']
    patrones = [quitar, poner, comma]
    return patrones


# 1. EXTRACT

def extract() -> list[pd.DataFrame]:
    file_names = ['data_dictionary.csv', 'order_details.csv', 'orders.csv',
                  'pizza_types.csv', 'pizzas.csv']
    df_lst = []
    for name in file_names:
        if name in ['data_dictionary.csv', 'pizzas.csv', 'pizza_types.csv']:
            sep = ','
        else:
            sep = ';'
        df = pd.read_csv(f'files2016/{name}', sep, encoding='latin_1')
        df_lst.append(df)
    return df_lst


# 2. TRANSFORM

def drop_nans(df_orders: pd.DataFrame, df_order_details: pd.DataFrame):
    """
    Dropeamos los NaNs de ambos dataframes. Intersecamos ambos dataframes
    para droppear lo que hemos sacado de un dataframe en el otro
    """
    df_order_details.dropna(inplace=True)
    or_id_A = set(df_orders['order_id'].unique())
    or_id_B = set(df_order_details['order_id'].unique())
    keep_order_id = or_id_A & or_id_B
    df_orders = df_orders[df_orders['order_id'].isin(keep_order_id)]
    df_order_details = df_order_details[df_order_details['order_id'].isin(keep_order_id)]
    # We sort the dataframe and reset the indexes
    df_orders.sort_values(by='order_id', inplace=True)
    df_orders.reset_index(drop=True, inplace=True)
    df_order_details.sort_values(by='order_id', inplace=True)
    df_order_details.reset_index(drop=True, inplace=True)
    return df_orders, df_order_details


def transform_key(key):
    if key[-1] == 's':
        end_str, count = 2, 0.75
    elif key[-1] == 'm':
        end_str, count = 2, 1
    elif key[-1] == 'l' and key[-2] != 'x':
        end_str, count = 2, 1.5
    elif key[-2:] == 'xl' and key[-3] != 'x':
        end_str, count = 3, 2
    else:
        end_str, count = 4, 3
    return end_str, count


def clean_data(df_orders: pd.DataFrame, df_order_details: pd.DataFrame):
    df_orders.dropna(subset='date', inplace=True)
    df_orders.reset_index(drop=True, inplace=True)
    df_orders.drop('time', axis=1, inplace=True)    
    
    # DATA CLEANING
    #   1. DATETIME FORMAT
    for i in range(len(df_orders)):
        unformatted_date = str(df_orders['date'][i])
        df_orders.loc[i, 'date'] = pd.to_datetime(df_orders['date'][i], errors='coerce')
        if pd.isnull(df_orders.loc[i, 'date']):
            unformatted_date = unformatted_date[:unformatted_date.find('.')]
            formatted_date = datetime.datetime.fromtimestamp(int(unformatted_date))
            df_orders.loc[i, 'date'] = pd.to_datetime(formatted_date)

    df_orders['date'] = pd.to_datetime(df_orders['date'], format="%Y/%m/%d")
    df_orders['week'] = df_orders['date'].dt.week
    df_orders['weekday'] = df_orders['date'].dt.weekday

    #   2. CORRECT NAMES
    df_orders, df_order_details = drop_nans(df_orders, df_order_details)
    patrones = compilar_patrones()
    [quitar, poner, comma] = patrones
    # Using regex, we correct some names and numbers from pizzas
    for i in range(len(quitar[:-2])):
        df_order_details['pizza_id'] = [quitar[i].sub(poner[i], str(x)) for x in df_order_details['pizza_id']]
    for i in range(len(quitar[:-2]), len(quitar)):
        df_order_details['quantity'] = [quitar[i].sub(poner[i], str(x)) for x in df_order_details['quantity']]
    df_order_details['quantity'] = [abs(int(x)) for x in df_order_details['quantity']]

    return df_orders, df_order_details


def get_pizzas_year(df_lst: list[pd.DataFrame], df_order_details: pd.DataFrame):
    df_pizzas = df_lst[4]

    # We create our dictionary of yearly pizzas
    pizzas_anual_dict = {}
    pizzas_anual_sizes_dict = {}
    for index in range(len(df_pizzas)):
        pizzas_anual_dict[df_pizzas['pizza_type_id'][index]] = 0
        pizzas_anual_sizes_dict[df_pizzas['pizza_id'][index]] = 0

    for index in range(1, len(df_order_details)):
        # With this method we access a specific value: df.iloc[columna].iloc[fila]
        key = df_order_details['pizza_id'].iloc[index]
        end_str, count = transform_key(key)
        value = df_order_details['quantity'].iloc[index]*count
        pizzas_anual_dict[key[:-end_str]] += value

    datos_anual = {'pizzas': list(pizzas_anual_dict.keys()),
                   'quantity': list(pizzas_anual_dict.values())}
    df_pizzas_anuales = pd.DataFrame.from_dict(datos_anual).round(decimals=0)
    df_pizzas_anuales.sort_values(by='quantity', inplace=True, ascending=False)
    df_pizzas_anuales.set_index('pizzas', inplace=True)

    return df_pizzas_anuales


def get_pizzas_weeks(df_orders, df_order_details, df_pizzas):
    # DATAFRAME FOR PIZZAS ORDERED EACH WEEK
    pizzas_anual_dict = {}
    pizzas_anual_sizes_dict = {}
    for index in range(len(df_pizzas)):
        pizzas_anual_dict[df_pizzas['pizza_type_id'][index]] = 0
        pizzas_anual_sizes_dict[df_pizzas['pizza_id'][index]] = 0

    weeks_dict = {}
    for week in range(1,52):
        # We create our dictionary of weekly pizzas
        pizzas_semana_dict = {}
        for pizza in pizzas_anual_dict:
            pizzas_semana_dict[pizza] = 0

        week_orders = df_orders.loc[df_orders['week']==week]
        order_ids = week_orders['order_id']
        week_df = df_order_details.loc[df_order_details['order_id'].isin(order_ids)]
        
        for index in range(1,len(week_df)):           
            key = week_df['pizza_id'].iloc[index]
            end_str, count = transform_key(key)
            value = week_df['quantity'].iloc[index]*count
            pizzas_semana_dict[key[:-end_str]] += value
            pizzas_anual_dict[key[:-end_str]] += value
        
        weeks_dict[week] = pizzas_semana_dict

        order_ids = week_orders['order_id']
        week_df = df_order_details.loc[df_order_details['order_id'].isin(order_ids)]
    
    df_weeks = pd.DataFrame.from_dict(weeks_dict).round(decimals=0)

    # DATAFRAME FOR PIZZAS ORDERED EACH WEEKDAY
    weekdays_dict = {}
    dayofweek = ['monday','tuesday','wednesday','thursday','friday','saturday','sunday']
    for day in range(7):
        pizzas_weekday_dict = {}
        for pizza in pizzas_anual_dict:
            pizzas_weekday_dict[pizza] = 0
            
        weekday_orders = df_orders.loc[df_orders['weekday']==day]
        order_ids = weekday_orders['order_id']
        weekday_df = df_order_details.loc[df_order_details['order_id'].isin(order_ids)]
        
        for index in range(1,len(week_df)):
            key = weekday_df['pizza_id'].iloc[index]
            end_str, count = transform_key(key)
            value = weekday_df['quantity'].iloc[index]*count
            pizzas_weekday_dict[key[:-end_str]] += value
        weekdays_dict[dayofweek[day]] = pizzas_weekday_dict

    df_weekdays = pd.DataFrame.from_dict(weekdays_dict).round(decimals=0)

    return (df_weekdays, df_weeks)


def transform(df_lst: list[pd.DataFrame], semana) -> tuple[pd.DataFrame]:
    df_order_details, df_orders, df_pizzas= df_lst[1], df_lst[2], df_lst[4]

    (df_orders, df_order_details) = clean_data(df_orders, df_order_details)
    df_pizzas_year = get_pizzas_year(df_lst, df_order_details)
    (df_weekdays, df_weeks) = get_pizzas_weeks(df_orders, df_order_details, df_pizzas)

    return (df_orders, df_order_details, df_pizzas_year, df_weekdays, df_weeks)


# 3. LOAD

def load_data_excel(df_results_lst: tuple[pd.DataFrame], name):
    (df_orders, df_order_details, df_pizzas_year, df_weekdays, df_weeks) = df_results_lst
    write = pd.ExcelWriter(f'{name}.xlsx')
    
    df_pizzas_year.to_excel(write, sheet_name='Yearly Pizzas')
    # For weeks and Weekdays, we are just interested in the sum, as we want
    # to see the orders placed each week or each weekday, respectively
    df_weeks.sum().round().to_excel(write, sheet_name='Weeks Pizzas')
    df_weekdays.sum().round().to_excel(write, sheet_name='Weekdays Pizzas')

    write.save()


def barchart_excel(sheet, chart_data: dict, transpose: bool):
    min_col = sheet.min_column
    max_col = sheet.max_column
    min_fila = sheet.min_row
    max_fila = sheet.max_row

    values = Reference(sheet, min_col=max_col, max_col=max_col, min_row=min_fila+1, max_row=max_fila)
    cats = Reference(sheet, min_col=min_col, max_col=min_col, min_row=min_fila+1, max_row=max_fila)

    barchart = BarChart()
    if transpose:
        barchart.type = 'bar'
    barchart.add_data(values, titles_from_data=True)
    barchart.set_categories(cats)

    barchart.legend = None

    barchart.title = chart_data['Title']
    barchart.x_axis.title = chart_data['x_axis']
    barchart.y_axis.title = chart_data['y_axis']
    barchart.y_axis.scaling.min = 0

    barchart.height = chart_data['height']
    barchart.width = chart_data['width']

    sheet.add_chart(barchart, 'D2')


def piechart_excel(sheet , chart_data: dict):
    min_col = sheet.min_column
    max_col = sheet.max_column
    min_fila = sheet.min_row
    max_fila = sheet.max_row

    values = Reference(sheet, min_col=max_col, max_col=max_col, min_row=min_fila+1, max_row=max_fila)
    cats = Reference(sheet, min_col=min_col, max_col=min_col, min_row=min_fila+1, max_row=max_fila)

    piechart = PieChart()
    piechart.add_data(values, titles_from_data=True)
    piechart.set_categories(cats)

    piechart.title = chart_data['Title']

    piechart.height = chart_data['height']
    piechart.width = chart_data['width']

    sheet.add_chart(piechart, chart_data['cell'])
  

def linechart_excel(sheet, chart_data: dict):
    min_col = sheet.min_column
    max_col = sheet.max_column
    min_fila = sheet.min_row
    max_fila = sheet.max_row

    values = Reference(sheet, min_col=max_col, max_col=max_col, min_row=min_fila+1, max_row=max_fila)
    # cats = Reference(sheet, min_col=min_col, max_col=min_col, min_row=min_fila+1, max_row=max_fila)

    linechart = LineChart()
    linechart.add_data(values, titles_from_data=True)
    # linechart.set_categories(cats)

    linechart.legend = None

    linechart.title = chart_data['Title']
    linechart.x_axis.title = chart_data['x_axis']
    linechart.y_axis.title = chart_data['y_axis']

    linechart.height = chart_data['height']
    linechart.width = chart_data['width']

    sheet.add_chart(linechart, 'D2')


def load(dfs_for_plotting: tuple[pd.DataFrame], file_name: str):
    # We load the dataframes into different sheets inside the file
    load_data_excel(dfs_for_plotting, file_name)
    # We get the WorkBook with the file load-workbook from openpyxl
    wb = load_workbook(f'{file_name}.xlsx')

    # SHEET 1
    sh1 = wb['Yearly Pizzas']
    # We create a dictionary for the data in the charts
    chart_data = {
        'Title':'(Ponderated) Pizzas ordered in 2016',
        'x_axis': 'Pizzas',
        'y_axis': 'Quantity',
        'height': 18,
        'width': 15,
        'cell': 'D2'
    }
    barchart_excel(sh1, chart_data, transpose=True)
    chart_data['cell'] = 'M2'
    chart_data['height'] = 10
    chart_data['width'] = 10
    piechart_excel(sh1, chart_data)

    # SHEET 2
    sh2 = wb['Weeks Pizzas']
    chart_data = {
        'Title':'Time Evolution of Pizzas Ordered by Week',
        'x_axis': 'Week',
        'y_axis': 'Pizzas Ordered',
        'height': 15,
        'width': 24,
        'cell': 'D2'
    }
    linechart_excel(sh2, chart_data)

    # SHEET 3
    sh3 = wb['Weekdays Pizzas']
    chart_data = {
        'Title':'Pizzas Ordered by Weekday',
        'x_axis': 'WeekDay',
        'y_axis': 'Pizzas Ordered',
        'height': 15,
        'width': 24,
        'cell': 'D2'
    }
    barchart_excel(sh3, chart_data, transpose=False)

    # We finally save the WorkBook
    wb.save(f'{file_name}.xlsx')


if __name__ == "__main__":
    dfs = extract()
    dfs_for_plotting = transform(dfs, 25)
    
    # File name (without '.xlsx' extension)
    file_name = 'report_maven_pizzas_2016'
    load(dfs_for_plotting, file_name)
    